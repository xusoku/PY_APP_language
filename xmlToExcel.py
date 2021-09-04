# coding=utf-8
# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

from xml.dom.minidom import Document
from xml.dom.minidom import parse
import openpyxl
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding( "utf-8" )

def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print("Hi, {0}".format(name))  # Press ⌘F8 to toggle the breakpoint.


def get_attrvalue(node, attrname):
    return node.getAttribute(attrname) if node else ''


def get_nodevalue(node, index=0):
    return node.childNodes[index].nodeValue if node else ''


def get_xmlnode(node, name):
    return node.getElementsByTagName(name) if node else []

def set_row(doc,key,value):
    row = doc.createElement('string')
    row.setAttribute('name', key)
    value = doc.createTextNode(value)
    row.appendChild(value)
    return row

def writeXml():
    wb = load_workbook("sample.xlsx")
    ws = wb.active
    writeFile(ws,1,"test_zh.xml")
    writeFile(ws,2,"test_en.xml")
    writeFile(ws,3,"test_es.xml")

def writeFile(ws,i,filename):  # i  1 中文 2 英文 3 西班牙
    doc = Document()
    root = doc.createElement('resources')  # 创建根元素
    root.setAttribute('xmlns:tools', "http://schemas.android.com/tools")  # 设置命名空间
    doc.appendChild(root)
    for row in ws:
        name = row[0].value
        value = row[i].value
        if not value:
            value=" "
        child = set_row(doc,name,value)
        root.appendChild(child)
    f = open(filename, 'w')
    f.write(doc.toprettyxml())
    f.close()


def readXml():
    zhXml = parse("strings_zh.xml")
    zhRoot = zhXml.documentElement
    zhNodes = get_xmlnode(zhRoot, 'string')
    zhMap = dict()

    enXml = parse("strings_en.xml")
    enRoot = enXml.documentElement
    enNodes = get_xmlnode(enRoot, 'string')
    enMap = dict()

    esXml = parse("strings_es.xml")
    esRoot = esXml.documentElement
    esNodes = get_xmlnode(esRoot, 'string')
    esMap = dict()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "name(key)"
    ws['B1'] = "Zh"
    ws['C1'] = "En"
    ws['D1'] = "Es"

    x = 0
    y = 0
    z = 0

    for esChild in esNodes:
        name = get_attrvalue(esChild, "name")
        value = get_nodevalue(esChild)
        esMap[name] = value
        y += 1
    for zhChild in zhNodes:
        name = get_attrvalue(zhChild, "name")
        value = get_nodevalue(zhChild)
        zhMap[name] = value
        z += 1

    for enChild in enNodes:
        name = get_attrvalue(enChild, "name")
        print name
        value = get_nodevalue(enChild)
        enMap[name] = value
        x += 1
        if (zhMap.has_key(name)):
            zhValue = zhMap.get(name)
        else:
            zhValue = ""

        if (esMap.has_key(name)):
            esValue = esMap.get(name)
        else:
            esValue = ""
        ws.append([name, zhValue, value, esValue])

    max = (x if (x > y) else y) if ((x if (x > y) else y) > z) else z
    print ("x=" + str(x) + "  y=" + str(y) + "  z=" + str(z) + "  max=" + str(max))

    wb.save("sample.xlsx")



if __name__ == '__main__':
    print_hi('PyCharm')
    writeXml()

